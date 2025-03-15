import csv
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from typing import List, Dict
from decimal import Decimal
from datetime import datetime
import io
import logging
from app.models import Invoice
from app.config import settings

logger = logging.getLogger(__name__)

class InvoiceExporter:
    def __init__(self):
        self.csv_columns = [
            "Filename", "Invoice Number", "Vendor Name", "Vendor Address",
            "Invoice Date", "Grand Total", "Taxes", "Final Total", "Pages",
            "Item Descriptions", "Item Quantities", "Item Unit Prices", "Item Totals"
        ]

    def export_to_csv(self, invoices: List[Invoice]) -> io.StringIO:
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=self.csv_columns)
        writer.writeheader()

        for invoice in invoices:
            row = self._prepare_csv_row(invoice)
            writer.writerow(row)

        output.seek(0)
        return output

    def export_to_excel(self, invoices: List[Invoice]) -> io.BytesIO:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Invoices"

        self._write_excel_header(sheet)

        for row, invoice in enumerate(invoices, start=2):
            self._write_excel_row(sheet, row, invoice)

        self._apply_excel_styling(sheet)

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output

    def _prepare_csv_row(self, invoice: Invoice) -> Dict:
        return {
            "Filename": invoice.filename,
            "Invoice Number": invoice.invoice_number,
            "Vendor Name": invoice.vendor.name,
            "Vendor Address": self._format_address(invoice.vendor.address),
            "Invoice Date": invoice.invoice_date.strftime("%Y-%m-%d"),
            "Grand Total": str(invoice.grand_total),
            "Taxes": str(invoice.taxes),
            "Final Total": str(invoice.final_total),
            "Pages": str(invoice.pages),
            "Item Descriptions": "|".join([item.description for item in invoice.items]),
            "Item Quantities": "|".join([str(item.quantity) for item in invoice.items]),
            "Item Unit Prices": "|".join([str(item.unit_price) for item in invoice.items]),
            "Item Totals": "|".join([str(item.total) for item in invoice.items])
        }

    def _format_address(self, address: Dict) -> str:
        return f"{address['street']}, {address['city']}, {address['state']} {address['postal_code']}, {address['country']}"

    def _write_excel_header(self, sheet: openpyxl.worksheet.worksheet.Worksheet):
        for col, header in enumerate(self.csv_columns, start=1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def _write_excel_row(self, sheet: openpyxl.worksheet.worksheet.Worksheet, row: int, invoice: Invoice):
        data = self._prepare_csv_row(invoice)
        for col, key in enumerate(self.csv_columns, start=1):
            sheet.cell(row=row, column=col, value=data[key])

    def _apply_excel_styling(self, sheet: openpyxl.worksheet.worksheet.Worksheet):
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                     top=Side(style='thin'), bottom=Side(style='thin'))
                if isinstance(cell.value, (int, float, Decimal)):
                    cell.number_format = '#,##0.00'
                elif isinstance(cell.value, datetime):
                    cell.number_format = 'YYYY-MM-DD'

        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width

def export_invoices(invoices: List[Invoice], format: str) -> io.BytesIO:
    exporter = InvoiceExporter()
    try:
        if format.lower() == 'csv':
            output = exporter.export_to_csv(invoices)
            return io.BytesIO(output.getvalue().encode())
        elif format.lower() == 'excel':
            return exporter.export_to_excel(invoices)
        else:
            raise ValueError(f"Unsupported export format: {format}")
    except Exception as e:
        logger.error(f"Error during invoice export: {str(e)}")
        raise

