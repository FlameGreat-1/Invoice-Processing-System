import gradio as gr
import requests
import json
import time
import os
from fastapi import FastAPI
from app.main import app as fastapi_app
from app.config import settings

# Initialize the FastAPI app
app = FastAPI()

# Mount the main FastAPI app
app.mount("/api", fastapi_app)

# Hugging Face Spaces configuration
HF_SPACE_URL = settings.HF_SPACE_URL

def process_invoices(files):
    try:
        # Upload files
        upload_url = f"{HF_SPACE_URL}/api/upload/"
        files_dict = [("files", (file.name, file.read(), file.type)) for file in files]
        response = requests.post(upload_url, files=files_dict)
        response.raise_for_status()
        
        task_id = response.json()["task_id"]
        
        # Poll for status
        status_url = f"{HF_SPACE_URL}/api/status/{task_id}"
        while True:
            status_response = requests.get(status_url)
            status_response.raise_for_status()
            
            status_data = status_response.json()
            if status_data["status"]["status"] == "Completed":
                break
            elif status_data["status"]["status"] == "Failed":
                return f"Processing failed: {status_data['status']['message']}"
            
            time.sleep(5)  # Wait 5 seconds before checking again
        
        # Download results
        csv_url = f"{HF_SPACE_URL}/api/download/{task_id}?format=csv"
        excel_url = f"{HF_SPACE_URL}/api/download/{task_id}?format=excel"
        
        csv_response = requests.get(csv_url)
        excel_response = requests.get(excel_url)
        
        csv_response.raise_for_status()
        excel_response.raise_for_status()
        
        # Save downloaded files
        output_dir = "output"
        os.makedirs(output_dir, exist_ok=True)
        csv_path = os.path.join(output_dir, f"invoices_{task_id}.csv")
        excel_path = os.path.join(output_dir, f"invoices_{task_id}.xlsx")
        
        with open(csv_path, "wb") as f:
            f.write(csv_response.content)
        with open(excel_path, "wb") as f:
            f.write(excel_response.content)
        
        return f"Processing completed. Results saved as {csv_path} and {excel_path}"
    except requests.RequestException as e:
        return f"Error during processing: {str(e)}"
    except Exception as e:
        return f"Unexpected error: {str(e)}"

# Define the Gradio interface
iface = gr.Interface(
    fn=process_invoices,
    inputs=gr.File(file_count="multiple", label="Upload Invoice Files (PDF, JPG, PNG, or ZIP)"),
    outputs="text",
    title=settings.PROJECT_NAME,
    description="Upload invoice files to extract and validate information. Results will be provided in CSV and Excel formats.",
)

# Combine FastAPI and Gradio
app = gr.mount_gradio_app(app, iface, path="/")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=settings.PORT)         # Use an official Python runtime as a parent image
FROM python:3.9-slim

# Set environment variables
ENV PYTHONDONTWRITEBYTECODE 1
ENV PYTHONUNBUFFERED 1
ENV DEBIAN_FRONTEND=noninteractive

# Set the working directory in the container
WORKDIR /app

# Install system dependencies
RUN apt-get update && apt-get install -y --no-install-recommends \
    build-essential \
    libpq-dev \
    tesseract-ocr \
    libtesseract-dev \
    poppler-utils \
    libmagic1 \
    && rm -rf /var/lib/apt/lists/*

# Install Python dependencies
COPY requirements.txt .
RUN pip install --no-cache-dir -r requirements.txt

# Download spaCy model
RUN python -m spacy download en_core_web_sm

# Copy the current directory contents into the container
COPY . .

# Make port 8000 available to the world outside this container
EXPOSE 8000

# Create a non-root user and switch to it
RUN adduser --disabled-password --gecos '' appuser
USER appuser

# Run the application
CMD ["gunicorn", "-k", "uvicorn.workers.UvicornWorker", "-b", "0.0.0.0:8000", "app:app"]                    # FastAPI and related
fastapi==0.68.1
uvicorn==0.15.0
python-multipart==0.0.5
pydantic==1.8.2
celery==5.1.2
redis==3.5.3

# Data processing and analysis
numpy==1.21.2
pandas==1.3.3

# OCR and image processing
Pillow==8.3.2
pytesseract==0.3.8
opencv-python-headless==4.5.3.56
pdf2image==1.16.0

# NLP and text processing
spacy==3.1.3
en_core_web_sm @ https://github.com/explosion/spacy-models/releases/download/en_core_web_sm-3.1.0/en_core_web_sm-3.1.0.tar.gz
dateparser==1.0.0
price-parser==0.3.4
usaddress==0.5.10
pycountry==20.7.3

# Machine learning and deep learning
torch==1.9.1
transformers==4.11.3
layoutlm==0.0.1

# Google Cloud Vision
google-cloud-vision==2.4.2

# File handling and compression
python-magic==0.4.24
zipfile36==0.1.3

# Excel handling
openpyxl==3.0.9
XlsxWriter==3.0.2

# Hugging Face and Gradio
gradio==2.4.0
huggingface-hub==0.0.19

# Asynchronous programming
aiohttp==3.7.4.post0
async-timeout==3.0.1

# Logging and error tracking
sentry-sdk==1.4.3

# Testing
pytest==6.2.5
pytest-asyncio==0.15.1

# Development tools
black==21.9b0
flake8==3.9.2
isort==5.9.3

# Production deployment
gunicorn==20.1.0                  from celery import Celery
from app.config import settings
from app.utils.file_handler import FileHandler
from app.utils.ocr_engine import ocr_engine
from app.utils.data_extractor import data_extractor
from app.utils.validator import invoice_validator, flag_anomalies
from app.utils.exporter import export_invoices
import os
import tempfile
from typing import List
import asyncio

celery_app = Celery('invoice_processing', broker=settings.CELERY_BROKER_URL)

file_handler = FileHandler()

@celery_app.task(bind=True)
def process_file_task(self, task_id: str, file_path: str):
    try:
        self.update_state(state='STARTED', meta={'progress': 0, 'message': 'Starting processing'})
        
        processed_files = file_handler.process_upload(file_path)
        self.update_state(state='STARTED', meta={'progress': 30, 'message': 'File processed'})

        loop = asyncio.get_event_loop()
        ocr_results = loop.run_until_complete(ocr_engine.process_images(processed_files))
        self.update_state(state='STARTED', meta={'progress': 60, 'message': 'OCR completed'})

        extracted_data = [data_extractor.extract_data(result) for result in ocr_results.values()]
        self.update_state(state='STARTED', meta={'progress': 80, 'message': 'Data extraction completed'})

        validated_data = [invoice for invoice, is_valid, _ in invoice_validator.validate_invoice_batch(extracted_data) if is_valid]
        flagged_invoices = flag_anomalies(validated_data)
        
        csv_output = export_invoices(validated_data, 'csv')
        excel_output = export_invoices(validated_data, 'excel')
        
        temp_dir = tempfile.gettempdir()
        csv_path = os.path.join(temp_dir, f"{task_id}_invoices.csv")
        excel_path = os.path.join(temp_dir, f"{task_id}_invoices.xlsx")
        
        with open(csv_path, 'wb') as f:
            f.write(csv_output.getvalue())
        with open(excel_path, 'wb') as f:
            f.write(excel_output.getvalue())
        
        self.update_state(state='SUCCESS', meta={'progress': 100, 'message': 'Processing completed'})
        
    except Exception as e:
        self.update_state(state='FAILURE', meta={'progress': 100, 'message': f'Error: {str(e)}'})
        raise

@celery_app.task(bind=True)
def process_multiple_files_task(self, task_id: str, file_paths: List[str]):
    try:
        self.update_state(state='STARTED', meta={'progress': 0, 'message': 'Starting processing'})
        
        processed_files = []
        for idx, file_path in enumerate(file_paths):
            processed_files.extend(file_handler.process_upload(file_path))
            progress = (idx + 1) / len(file_paths) * 30
            self.update_state(state='STARTED', meta={'progress': progress, 'message': f'Processed {idx + 1} of {len(file_paths)} files'})

        loop = asyncio.get_event_loop()
        ocr_results = loop.run_until_complete(ocr_engine.process_images(processed_files))
        self.update_state(state='STARTED', meta={'progress': 60, 'message': 'OCR completed'})

        extracted_data = [data_extractor.extract_data(result) for result in ocr_results.values()]
        self.update_state(state='STARTED', meta={'progress': 80, 'message': 'Data extraction completed'})

        validated_data = [invoice for invoice, is_valid, _ in invoice_validator.validate_invoice_batch(extracted_data) if is_valid]
        flagged_invoices = flag_anomalies(validated_data)
        
        csv_output = export_invoices(validated_data, 'csv')
        excel_output = export_invoices(validated_data, 'excel')
        
        temp_dir = tempfile.gettempdir()
        csv_path = os.path.join(temp_dir, f"{task_id}_invoices.csv")
        excel_path = os.path.join(temp_dir, f"{task_id}_invoices.xlsx")
        
        with open(csv_path, 'wb') as f:
            f.write(csv_output.getvalue())
        with open(excel_path, 'wb') as f:
            f.write(excel_output.getvalue())
        
        self.update_state(state='SUCCESS', meta={'progress': 100, 'message': 'Processing completed'})
        
    except Exception as e:
        self.update_state(state='FAILURE', meta={'progress': 100, 'message': f'Error: {str(e)}'})
        raise

celery_app.conf.task_routes = {
    'app.celery_app.process_file_task': {'queue': 'single_file'},
    'app.celery_app.process_multiple_files_task': {'queue': 'multiple_files'},
}                  # app/config.py

import os
from pydantic import BaseSettings, Field

class Settings(BaseSettings):
    # API Configuration
    API_V1_STR: str = "/api/v1"
    PROJECT_NAME: str = "Invoice Processing System"

    # File Upload Configuration
    MAX_UPLOAD_SIZE: int = 100 * 1024 * 1024  # 100MB
    ALLOWED_EXTENSIONS: set = {"pdf", "jpg", "jpeg", "png", "zip"}

    # OCR Configuration
    OCR_MODEL_ID: str = "jinhybr/OCR-LayoutLMv3-Invoice"
    OCR_TIMEOUT: int = 300  # 5 minutes

    # Processing Configuration
    MULTI_PAGE_THRESHOLD: float = 0.95  # 95% confidence for multi-page detection
    INVOICE_NUMBER_ACCURACY: float = 0.95  # 95% accuracy for invoice number extraction
    TOTAL_MATH_ACCURACY: float = 1.0  # 100% accuracy for total calculations
    MAX_WORKERS: int = 5  # or any other appropriate number

    # Output Configuration
    OUTPUT_FORMAT: str = "csv"  # or "excel"

    # Hugging Face Configuration
    HF_API_TOKEN: str = Field(..., env="HF_API_TOKEN")

    # Google Cloud Vision Configuration (Backup OCR)
    GCV_CREDENTIALS: str = Field(..., env="GOOGLE_APPLICATION_CREDENTIALS")

    # Database Configuration (for potential future use)
    DATABASE_URL: str = Field(..., env="DATABASE_URL")

    class Config:
        env_file = ".env"
        case_sensitive = True

settings = Settings()

def get_settings() -> Settings:
    return settings
    from fastapi import FastAPI, File, UploadFile, HTTPException, BackgroundTasks
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import List, Optional
import asyncio
import tempfile
import os
import uuid
from app.config import settings
from app.utils.file_handler import FileHandler
from app.utils.ocr_engine import ocr_engine
from app.utils.data_extractor import data_extractor
from app.utils.validator import invoice_validator, flag_anomalies
from app.utils.exporter import export_invoices
from app.models import Invoice, ProcessingStatus
from app.celery_app import process_file_task, process_multiple_files_task
from celery.result import AsyncResult
import logging

app = FastAPI(title=settings.PROJECT_NAME, version="1.0.0")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

file_handler = FileHandler()

class ProcessingRequest(BaseModel):
    task_id: str

class ProcessingResponse(BaseModel):
    task_id: str
    status: ProcessingStatus

processing_tasks = {}

@app.post("/upload/", response_model=ProcessingRequest)
async def upload_files(files: List[UploadFile] = File(...)):
    task_id = str(uuid.uuid4())
    processing_tasks[task_id] = ProcessingStatus(status="Queued", progress=0, message="Task queued")
    
    temp_dir = tempfile.mkdtemp()
    file_paths = []

    for file in files:
        file_path = os.path.join(temp_dir, file.filename)
        with open(file_path, "wb") as buffer:
            content = await file.read()
            buffer.write(content)
        file_paths.append(file_path)

    if len(files) == 1:
        # Single file processing
        celery_task = process_file_task.delay(task_id, file_paths[0])
    else:
        # Multiple files processing
        celery_task = process_multiple_files_task.delay(task_id, file_paths)
    
    processing_tasks[task_id] = ProcessingStatus(status="Processing", progress=0, message="Processing started")
    
    return ProcessingRequest(task_id=task_id)

@app.get("/status/{task_id}", response_model=ProcessingResponse)
async def get_processing_status(task_id: str):
    if task_id not in processing_tasks:
        raise HTTPException(status_code=404, detail="Task not found")
    
    celery_task = AsyncResult(task_id)
    if celery_task.state == 'PENDING':
        status = "Queued"
    elif celery_task.state == 'STARTED':
        status = "Processing"
    elif celery_task.state == 'SUCCESS':
        status = "Completed"
    else:
        status = "Failed"
    
    progress = celery_task.info.get('progress', 0) if celery_task.info else 0
    message = celery_task.info.get('message', '') if celery_task.info else ''
    
    return ProcessingResponse(task_id=task_id, status=ProcessingStatus(status=status, progress=progress, message=message))

@app.get("/download/{task_id}")
async def download_results(task_id: str, format: str = "csv"):
    if task_id not in processing_tasks:
        raise HTTPException(status_code=404, detail="Task not found")
    
    celery_task = AsyncResult(task_id)
    if celery_task.state != 'SUCCESS':
        raise HTTPException(status_code=400, detail="Processing not completed")
    
    temp_dir = tempfile.gettempdir()
    if format.lower() == "csv":
        file_path = os.path.join(temp_dir, f"{task_id}_invoices.csv")
        media_type = "text/csv"
    elif format.lower() == "excel":
        file_path = os.path.join(temp_dir, f"{task_id}_invoices.xlsx")
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        raise HTTPException(status_code=400, detail="Invalid format specified")
    
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Result file not found")
    
    return FileResponse(file_path, media_type=media_type, filename=os.path.basename(file_path))

@app.on_event("startup")
async def startup_event():
    logger.info("Application is starting up")

@app.on_event("shutdown")
async def shutdown_event():
    logger.info("Application is shutting down")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)                             from pydantic import BaseModel, Field, validator
from typing import List, Optional
from datetime import date
from decimal import Decimal
import re

class Address(BaseModel):
    street: str
    city: str
    state: Optional[str]
    country: str
    postal_code: str

class Vendor(BaseModel):
    name: str
    address: Address

class InvoiceItem(BaseModel):
    description: str
    quantity: int
    unit_price: Decimal
    total: Decimal

class Invoice(BaseModel):
    filename: str
    invoice_number: str = Field(..., regex=r'^INV-\d{3}$')
    vendor: Vendor
    invoice_date: date
    grand_total: Decimal
    taxes: Decimal
    final_total: Decimal
    items: List[InvoiceItem]
    pages: int = Field(ge=1)

    @validator('final_total')
    def validate_final_total(cls, v, values):
        if 'grand_total' in values and 'taxes' in values:
            expected_total = values['grand_total'] + values['taxes']
            if abs(v - expected_total) > Decimal('0.01'):
                raise ValueError(f"Final total {v} does not match grand total {values['grand_total']} plus taxes {values['taxes']}")
        return v

    @validator('invoice_date')
    def validate_invoice_date(cls, v):
        if v > date.today():
            raise ValueError("Invoice date cannot be in the future")
        return v

class ProcessingResult(BaseModel):
    success: bool
    message: str
    invoices: List[Invoice] = []
    errors: List[str] = []

class FileUpload(BaseModel):
    filename: str
    content_type: str
    file_size: int

    @validator('content_type')
    def validate_content_type(cls, v):
        allowed_types = {'application/pdf', 'image/jpeg', 'image/png', 'application/zip'}
        if v not in allowed_types:
            raise ValueError(f"Unsupported file type: {v}")
        return v

    @validator('file_size')
    def validate_file_size(cls, v):
        max_size = 100 * 1024 * 1024  # 100MB
        if v > max_size:
            raise ValueError(f"File size exceeds maximum allowed size of 100MB")
        return v

class ExportFormat(BaseModel):
    format: str = Field(..., regex='^(csv|excel)$')

class ProcessingStatus(BaseModel):
    status: str
    progress: float = Field(ge=0, le=100)
    message: Optional[str]
    import re
from typing import Dict, List, Tuple
from datetime import datetime
from decimal import Decimal
import spacy
from spacy.matcher import Matcher
from spacy.tokens import Doc
from app.models import Invoice, Vendor, Address, InvoiceItem
from app.config import settings
import usaddress
import pycountry
import dateparser
from price_parser import Price
import asyncio
from concurrent.futures import ThreadPoolExecutor

class DataExtractor:
    def __init__(self):
        self.nlp = spacy.load("en_core_web_sm")
        self.matcher = Matcher(self.nlp.vocab)
        self._setup_matchers()
        self.executor = ThreadPoolExecutor(max_workers=settings.MAX_WORKERS)

    def _setup_matchers(self):
        self.matcher.add("INVOICE_NUMBER", [[{"LOWER": "invoice"}, {"LOWER": "number"}, {"IS_ASCII": True, "LENGTH": {">=": 5}}]])
        self.matcher.add("INVOICE_DATE", [[{"LOWER": "date"}, {"IS_ASCII": True, "LENGTH": {">=": 8, "<=": 10}}]])
        self.matcher.add("TOTAL_AMOUNT", [[{"LOWER": {"IN": ["total", "amount", "sum"]}}, {"LIKE_NUM": True}]])
        self.matcher.add("TAX_AMOUNT", [[{"LOWER": {"IN": ["tax", "vat", "gst"]}}, {"LIKE_NUM": True}]])

    async def extract_data(self, ocr_result: Dict) -> Invoice:
        if ocr_result.get("is_multipage", False):
            return await self._extract_multipage_data(ocr_result)
        else:
            return await self._extract_single_page_data(ocr_result)

    async def _extract_multipage_data(self, ocr_result: Dict) -> Invoice:
        text = " ".join(ocr_result["words"])
        doc = await self._process_text(text)

        invoice_number = self._extract_invoice_number(doc)
        vendor = self._extract_vendor(doc)
        invoice_date = self._extract_date(doc)
        grand_total, taxes, final_total = self._extract_totals(doc)
        items = await self._extract_items_multipage(doc, ocr_result["boxes"])

        return Invoice(
            filename=ocr_result.get("filename", ""),
            invoice_number=invoice_number,
            vendor=vendor,
            invoice_date=invoice_date,
            grand_total=grand_total,
            taxes=taxes,
            final_total=final_total,
            items=items,
            pages=ocr_result.get("num_pages", 1)
        )

    async def _extract_single_page_data(self, ocr_result: Dict) -> Invoice:
        text = " ".join(ocr_result["words"])
        doc = await self._process_text(text)

        invoice_number = self._extract_invoice_number(doc)
        vendor = self._extract_vendor(doc)
        invoice_date = self._extract_date(doc)
        grand_total, taxes, final_total = self._extract_totals(doc)
        items = await self._extract_items(doc, ocr_result["boxes"])

        return Invoice(
            filename=ocr_result.get("filename", ""),
            invoice_number=invoice_number,
            vendor=vendor,
            invoice_date=invoice_date,
            grand_total=grand_total,
            taxes=taxes,
            final_total=final_total,
            items=items,
            pages=1
        )

    async def _process_text(self, text: str) -> Doc:
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(self.executor, self.nlp, text)

    def _extract_invoice_number(self, doc: Doc) -> str:
        matches = self.matcher(doc)
        for match_id, start, end in matches:
            if self.nlp.vocab.strings[match_id] == "INVOICE_NUMBER":
                return doc[end-1].text
        # Fallback: look for any alphanumeric string that looks like an invoice number
        for token in doc:
            if re.match(r'^[A-Za-z0-9-]{5,}$', token.text):
                return token.text
        return ""

    def _extract_vendor(self, doc: Doc) -> Vendor:
        org_names = [ent.text for ent in doc.ents if ent.label_ == "ORG"]
        if org_names:
            vendor_name = max(org_names, key=len)  # Choose the longest organization name
        else:
            vendor_name = "Unknown"
        
        address = self._extract_address(doc)
        return Vendor(name=vendor_name, address=address)

    def _extract_address(self, doc: Doc) -> Address:
        address_text = " ".join([ent.text for ent in doc.ents if ent.label_ in ["GPE", "LOC"]])
        parsed_address, _ = usaddress.tag(address_text)
        
        street = parsed_address.get('AddressNumber', '') + ' ' + parsed_address.get('StreetName', '')
        city = parsed_address.get('PlaceName', '')
        state = parsed_address.get('StateName', '')
        postal_code = parsed_address.get('ZipCode', '')
        
        country_name = parsed_address.get('CountryName', '')
        country = pycountry.countries.search_fuzzy(country_name)[0].alpha_2 if country_name else ''

        return Address(
            street=street.strip(),
            city=city,
            state=state,
            country=country,
            postal_code=postal_code
        )

    def _extract_date(self, doc: Doc) -> datetime:
        matches = self.matcher(doc)
        for match_id, start, end in matches:
            if self.nlp.vocab.strings[match_id] == "INVOICE_DATE":
                date_str = doc[end-1].text
                parsed_date = dateparser.parse(date_str)
                if parsed_date:
                    return parsed_date
        
        # Fallback: look for any date-like string
        for token in doc:
            parsed_date = dateparser.parse(token.text)
            if parsed_date:
                return parsed_date
        
        return datetime.now()  # Default to current date if not found

    def _extract_totals(self, doc: Doc) -> Tuple[Decimal, Decimal, Decimal]:
        grand_total = Decimal('0.00')
        taxes = Decimal('0.00')
        final_total = Decimal('0.00')

        matches = self.matcher(doc)
        for match_id, start, end in matches:
            if self.nlp.vocab.strings[match_id] == "TOTAL_AMOUNT":
                price = Price.fromstring(doc[end-1].text)
                if price.amount:
                    final_total = Decimal(str(price.amount))
            elif self.nlp.vocab.strings[match_id] == "TAX_AMOUNT":
                price = Price.fromstring(doc[end-1].text)
                if price.amount:
                    taxes = Decimal(str(price.amount))

        grand_total = final_total - taxes
        return grand_total, taxes, final_total

    async def _extract_items(self, doc: Doc, boxes: List[List[int]]) -> List[InvoiceItem]:
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(self.executor, self._extract_items_sync, doc, boxes)

    def _extract_items_sync(self, doc: Doc, boxes: List[List[int]]) -> List[InvoiceItem]:
        items = []
        for i, sent in enumerate(doc.sents):
            if any(token.like_num for token in sent):
                description = " ".join([token.text for token in sent if not token.like_num and not token.is_currency])
                numbers = [Price.fromstring(token.text) for token in sent if token.like_num or token.is_currency]
                if len(numbers) >= 3:
                    quantity = int(numbers[0].amount) if numbers[0].amount else 1
                    unit_price = Decimal(str(numbers[1].amount)) if numbers[1].amount else Decimal('0.00')
                    total = Decimal(str(numbers[2].amount)) if numbers[2].amount else Decimal('0.00')
                    
                    # Use bounding box information to determine if this is likely a line item
                    if i < len(boxes) and (boxes[i][2] - boxes[i][0]) > (doc[0].doc.page.width * 0.5):
                        items.append(InvoiceItem(
                            description=description,
                            quantity=quantity,
                            unit_price=unit_price,
                            total=total
                        ))
        return items

    async def _extract_items_multipage(self, doc: Doc, boxes: List[List[int]]) -> List[InvoiceItem]:
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(self.executor, self._extract_items_multipage_sync, doc, boxes)

    def _extract_items_multipage_sync(self, doc: Doc, boxes: List[List[int]]) -> List[InvoiceItem]:
        items = []
        item_start_phrases = ["item", "description", "product"]
        in_item_section = False
        current_item = None

        for i, sent in enumerate(doc.sents):
            sent_text = sent.text.lower()
            
            # Check if we're entering the item section
            if any(phrase in sent_text for phrase in item_start_phrases):
                in_item_section = True
                continue

            if in_item_section:
                if current_item is None:
                    current_item = {"description": "", "quantity": None, "unit_price": None, "total": None}

                # Check if this sentence contains numeric values
                numbers = [Price.fromstring(token.text) for token in sent if token.like_num or token.is_currency]
                
                if numbers:
                    # If we have numbers, try to fill in the item details
                    if len(numbers) >= 3:
                        current_item["quantity"] = int(numbers[0].amount) if numbers[0].amount else 1
                        current_item["unit_price"] = Decimal(str(numbers[1].amount)) if numbers[1].amount else Decimal('0.00')
                        current_item["total"] = Decimal(str(numbers[2].amount)) if numbers[2].amount else Decimal('0.00')
                        
                        # Add the item and reset
                        items.append(InvoiceItem(**current_item))
                        current_item = None
                else:
                    # If no numbers, add to the description
                    current_item["description"] += " " + sent.text

        # Add any remaining item
        if current_item and current_item["quantity"] is not None:
            items.append(InvoiceItem(**current_item))

        return items

data_extractor = DataExtractor()                import csv
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from typing import List, Dict
from decimal import Decimal
from datetime import datetime
import io
import logging
from app.models import Invoice, Address
from app.config import settings
import asyncio
from concurrent.futures import ThreadPoolExecutor

logger = logging.getLogger(__name__)

class InvoiceExporter:
    def __init__(self):
        self.csv_columns = [
            "Filename", "Invoice Number", "Vendor Name", "Vendor Address",
            "Invoice Date", "Grand Total", "Taxes", "Final Total", "Pages",
            "Item Descriptions", "Item Quantities", "Item Unit Prices", "Item Totals"
        ]
        self.executor = ThreadPoolExecutor(max_workers=settings.MAX_WORKERS)

    async def export_to_csv(self, invoices: List[Invoice]) -> io.StringIO:
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(self.executor, self._export_to_csv_sync, invoices)

    def _export_to_csv_sync(self, invoices: List[Invoice]) -> io.StringIO:
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=self.csv_columns)
        writer.writeheader()

        for invoice in invoices:
            row = self._prepare_csv_row(invoice)
            writer.writerow(row)

        output.seek(0)
        return output

    async def export_to_excel(self, invoices: List[Invoice]) -> io.BytesIO:
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(self.executor, self._export_to_excel_sync, invoices)

    def _export_to_excel_sync(self, invoices: List[Invoice]) -> io.BytesIO:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Invoices"

        self._write_excel_header(sheet)

        for row, invoice in enumerate(invoices, start=2):
            self._write_excel_row(sheet, row, invoice)

        self._apply_excel_styling(sheet)

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output

    def _prepare_csv_row(self, invoice: Invoice) -> Dict:
        return {
            "Filename": invoice.filename,
            "Invoice Number": invoice.invoice_number,
            "Vendor Name": invoice.vendor.name,
            "Vendor Address": self._format_address(invoice.vendor.address),
            "Invoice Date": invoice.invoice_date.strftime("%Y-%m-%d"),
            "Grand Total": str(invoice.grand_total),
            "Taxes": str(invoice.taxes),
            "Final Total": str(invoice.final_total),
            "Pages": str(invoice.pages),
            "Item Descriptions": "|".join([item.description for item in invoice.items]),
            "Item Quantities": "|".join([str(item.quantity) for item in invoice.items]),
            "Item Unit Prices": "|".join([str(item.unit_price) for item in invoice.items]),
            "Item Totals": "|".join([str(item.total) for item in invoice.items])
        }

    def _format_address(self, address: Address) -> str:
        return f"{address.street}, {address.city}, {address.state} {address.postal_code}, {address.country}"

    def _write_excel_header(self, sheet: openpyxl.worksheet.worksheet.Worksheet):
        for col, header in enumerate(self.csv_columns, start=1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def _write_excel_row(self, sheet: openpyxl.worksheet.worksheet.Worksheet, row: int, invoice: Invoice):
        data = self._prepare_csv_row(invoice)
        for col, key in enumerate(self.csv_columns, start=1):
            sheet.cell(row=row, column=col, value=data[key])

    def _apply_excel_styling(self, sheet: openpyxl.worksheet.worksheet.Worksheet):
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                     top=Side(style='thin'), bottom=Side(style='thin'))
                if isinstance(cell.value, (int, float, Decimal)):
                    cell.number_format = '#,##0.00'
                elif isinstance(cell.value, datetime):
                    cell.number_format = 'YYYY-MM-DD'

        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width

async def export_invoices(invoices: List[Invoice], format: str) -> io.BytesIO:
    exporter = InvoiceExporter()
    try:
        if format.lower() == 'csv':
            output = await exporter.export_to_csv(invoices)
            return io.BytesIO(output.getvalue().encode())
        elif format.lower() == 'excel':
            return await exporter.export_to_excel(invoices)
        else:
            raise ValueError(f"Unsupported export format: {format}")
    except Exception as e:
        logger.error(f"Error during invoice export: {str(e)}")
        raise                            import os
import zipfile
import magic
from fastapi import UploadFile, HTTPException
from typing import List, Dict
from PIL import Image
import fitz  # PyMuPDF
import io
import uuid
import asyncio
from concurrent.futures import ThreadPoolExecutor
from app.config import settings
from app.models import FileUpload

class FileHandler:
    def __init__(self, upload_dir: str = "/tmp/invoice_uploads"):
        self.upload_dir = upload_dir
        os.makedirs(self.upload_dir, exist_ok=True)
        self.executor = ThreadPoolExecutor(max_workers=settings.MAX_WORKERS)

    async def save_upload(self, file: UploadFile) -> FileUpload:
        content_type = magic.from_buffer(await file.read(1024), mime=True)
        await file.seek(0)
        
        if content_type not in settings.ALLOWED_EXTENSIONS:
            raise HTTPException(status_code=400, detail=f"Unsupported file type: {content_type}")
        
        file_size = 0
        file_path = os.path.join(self.upload_dir, f"{uuid.uuid4()}_{file.filename}")
        
        with open(file_path, "wb") as buffer:
            while chunk := await file.read(8192):
                file_size += len(chunk)
                if file_size > settings.MAX_UPLOAD_SIZE:
                    os.remove(file_path)
                    raise HTTPException(status_code=400, detail="File size exceeds the maximum allowed size")
                buffer.write(chunk)
        
        return FileUpload(filename=file_path, content_type=content_type, file_size=file_size)

    async def process_uploads(self, file_uploads: List[FileUpload]) -> List[Dict[str, any]]:
        tasks = [self.process_upload(file_upload) for file_upload in file_uploads]
        results = await asyncio.gather(*tasks)
        return [item for sublist in results for item in sublist]  # Flatten the list of lists

    async def process_upload(self, file_upload: FileUpload) -> List[Dict[str, any]]:
        if file_upload.content_type == 'application/zip':
            return await self._process_zip(file_upload.filename)
        elif file_upload.content_type == 'application/pdf':
            return await self._process_pdf(file_upload.filename)
        else:  # Image files
            return [await self._process_image(file_upload.filename)]

    async def _process_zip(self, zip_path: str) -> List[Dict[str, any]]:
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(self.executor, self._process_zip_sync, zip_path)

    def _process_zip_sync(self, zip_path: str) -> List[Dict[str, any]]:
        extracted_files = []
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            for file_info in zip_ref.infolist():
                if not file_info.filename.endswith('/'):  # Not a directory
                    with zip_ref.open(file_info) as file:
                        content = file.read()
                        content_type = magic.from_buffer(content, mime=True)
                        if content_type in settings.ALLOWED_EXTENSIONS:
                            if content_type == 'application/pdf':
                                extracted_files.extend(self._process_pdf_content(file_info.filename, content))
                            else:
                                extracted_files.append(self._process_image_content(file_info.filename, content))
        return extracted_files

    async def _process_pdf(self, pdf_path: str) -> List[Dict[str, any]]:
        loop = asyncio.get_event_loop()
        with open(pdf_path, 'rb') as file:
            content = await loop.run_in_executor(self.executor, file.read)
        return await loop.run_in_executor(self.executor, self._process_pdf_content, os.path.basename(pdf_path), content)

    def _process_pdf_content(self, filename: str, content: bytes) -> List[Dict[str, any]]:
        doc = fitz.open(stream=content, filetype="pdf")
        pages = []
        for page_num in range(len(doc)):
            page = doc.load_page(page_num)
            pix = page.get_pixmap()
            img_bytes = pix.tobytes("png")
            pages.append({
                'filename': f"{filename}_page_{page_num+1}.png",
                'content': img_bytes,
                'page_number': page_num + 1,
                'total_pages': len(doc)
            })
        doc.close()
        return [{
            'filename': filename,
            'content': content,
            'pages': pages,
            'is_multipage': len(pages) > 1
        }]

    async def _process_image(self, image_path: str) -> Dict[str, any]:
        loop = asyncio.get_event_loop()
        with open(image_path, 'rb') as file:
            content = await loop.run_in_executor(self.executor, file.read)
        return await loop.run_in_executor(self.executor, self._process_image_content, os.path.basename(image_path), content)

    def _process_image_content(self, filename: str, content: bytes) -> Dict[str, any]:
        return {
            'filename': filename,
            'content': content,
            'pages': [{
                'filename': filename,
                'content': content,
                'page_number': 1,
                'total_pages': 1
            }],
            'is_multipage': False
        }

    async def clean_up(self, file_path: str):
        loop = asyncio.get_event_loop()
        await loop.run_in_executor(self.executor, self._clean_up_sync, file_path)

    def _clean_up_sync(self, file_path: str):
        try:
            os.remove(file_path)
        except OSError as e:
            print(f"Error deleting file {file_path}: {e}")

file_handler = FileHandler()                                                           import asyncio
from typing import List, Tuple, Dict
import aiohttp
from transformers import LayoutLMv3FeatureExtractor, LayoutLMv3Tokenizer, LayoutLMv3ForSequenceClassification, LayoutLMv3Processor
import torch
from PIL import Image
import io
import numpy as np
from google.cloud import vision
from google.cloud.vision import types
from app.config import settings
from app.models import ProcessingStatus
import logging
import cv2
import pytesseract
import fitz  # PyMuPDF for PDF handling
from concurrent.futures import ThreadPoolExecutor

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class OCREngine:
    def __init__(self):
        self.model = LayoutLMv3ForSequenceClassification.from_pretrained(settings.OCR_MODEL_ID)
        self.feature_extractor = LayoutLMv3FeatureExtractor.from_pretrained(settings.OCR_MODEL_ID)
        self.tokenizer = LayoutLMv3Tokenizer.from_pretrained(settings.OCR_MODEL_ID)
        self.processor = LayoutLMv3Processor.from_pretrained(settings.OCR_MODEL_ID)
        self.device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
        self.model.to(self.device)
        self.gcv_client = vision.ImageAnnotatorClient()
        self.cache = {}  # Simple cache for storing processed results
        self.executor = ThreadPoolExecutor(max_workers=settings.MAX_WORKERS)

    async def process_documents(self, documents: List[Dict[str, any]]) -> Dict[str, Dict]:
        results = {}
        batches = self._create_batches(documents, batch_size=settings.BATCH_SIZE)
        
        for batch in batches:
            batch_results = await asyncio.gather(*[self._process_document(doc) for doc in batch])
            results.update(batch_results)
        
        return results

    def _create_batches(self, documents: List[Dict[str, any]], batch_size: int) -> List[List[Dict[str, any]]]:
        return [documents[i:i + batch_size] for i in range(0, len(documents), batch_size)]

    async def _process_document(self, document: Dict[str, any]) -> Tuple[str, Dict]:
        try:
            cache_key = hash(document['content'])
            if cache_key in self.cache:
                return document['filename'], self.cache[cache_key]

            if document['is_multipage']:
                ocr_result = await self._process_multipage_with_layoutlm(document)
            else:
                ocr_result = await self._process_single_page(document)

            self.cache[cache_key] = ocr_result
            return document['filename'], ocr_result
        except Exception as e:
            logger.error(f"Error processing {document['filename']}: {str(e)}")
            return document['filename'], {"error": str(e)}

    async def _process_multipage_with_layoutlm(self, document: Dict[str, any]) -> Dict:
        pages = [Image.open(io.BytesIO(page['content'])) for page in document['pages']]

        # Process all pages together
        encoding = self.processor(pages, return_tensors="pt", padding=True, truncation=True)
        for k, v in encoding.items():
            encoding[k] = v.to(self.device)

        with torch.no_grad():
            outputs = self.model(**encoding)

        # Extract text and bounding boxes for all pages
        words, boxes = [], []
        for page, page_encoding in zip(pages, encoding['input_ids']):
            page_words, page_boxes = self._extract_text_and_boxes_layoutlm(page, page_encoding)
            words.extend(page_words)
            boxes.extend(page_boxes)

        return {
            "words": words,
            "boxes": boxes,
            "is_multipage": True,
            "num_pages": len(pages)
        }

    async def _process_single_page(self, document: Dict[str, any]) -> Dict:
        image_bytes = document['content']
        image_name = document['filename']
        
        try:
            ocr_result = await self._process_with_layoutlm(image_name, image_bytes)
            if not ocr_result:
                logger.warning(f"LayoutLM failed for {image_name}, falling back to Google Cloud Vision")
                ocr_result = await self._process_with_gcv(image_name, image_bytes)
            if not ocr_result:
                logger.warning(f"GCV failed for {image_name}, falling back to Tesseract")
                ocr_result = await self._process_with_tesseract(image_name, image_bytes)
            return ocr_result
        except Exception as e:
            logger.error(f"Error in single page processing for {image_name}: {str(e)}")
            return {"error": str(e)}

    async def _process_with_layoutlm(self, image_name: str, image_bytes: bytes) -> Dict:
        image = Image.open(io.BytesIO(image_bytes)).convert("RGB")
        encoding = self.processor(image, return_tensors="pt")
        for k, v in encoding.items():
            encoding[k] = v.to(self.device)
        
        with torch.no_grad():
            outputs = self.model(**encoding)
        
        words, boxes = self._extract_text_and_boxes_layoutlm(image, encoding['input_ids'][0])
        
        return {
            "words": words,
            "boxes": boxes,
            "is_multipage": False,
            "num_pages": 1
        }

    def _extract_text_and_boxes_layoutlm(self, image: Image, encoding) -> Tuple[List[str], List[List[int]]]:
        words = self.tokenizer.convert_ids_to_tokens(encoding)
        words = [word for word in words if word not in [self.tokenizer.pad_token, self.tokenizer.cls_token, self.tokenizer.sep_token]]
        
        boxes = self.feature_extractor(image)['bbox'][0]
        boxes = [box for box, word in zip(boxes, words) if word not in [self.tokenizer.pad_token, self.tokenizer.cls_token, self.tokenizer.sep_token]]

        return words, boxes

    async def _process_with_gcv(self, image_name: str, image_bytes: bytes) -> Dict:
        image = types.Image(content=image_bytes)
        response = await asyncio.to_thread(self.gcv_client.document_text_detection, image)
        document = response.full_text_annotation

        words = []
        boxes = []
        for page in document.pages:
            for block in page.blocks:
                for paragraph in block.paragraphs:
                    for word in paragraph.words:
                        words.append(''.join([symbol.text for symbol in word.symbols]))
                        vertices = [(vertex.x, vertex.y) for vertex in word.bounding_box.vertices]
                        boxes.append(vertices)

        return {
            "words": words,
            "boxes": boxes,
            "is_multipage": False,
            "num_pages": 1
        }

    async def _process_with_tesseract(self, image_name: str, image_bytes: bytes) -> Dict:
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(self.executor, self._process_with_tesseract_sync, image_name, image_bytes)

    def _process_with_tesseract_sync(self, image_name: str, image_bytes: bytes) -> Dict:
        image = Image.open(io.BytesIO(image_bytes))
        image_np = np.array(image)

        gray = cv2.cvtColor(image_np, cv2.COLOR_RGB2GRAY)
        threshold = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]

        custom_config = r'--oem 3 --psm 6'
        details = pytesseract.image_to_data(threshold, output_type=pytesseract.Output.DICT, config=custom_config, lang='eng')

        words = []
        boxes = []
        n_boxes = len(details['text'])
        for i in range(n_boxes):
            if int(details['conf'][i]) > 60:
                (x, y, w, h) = (details['left'][i], details['top'][i], details['width'][i], details['height'][i])
                words.append(details['text'][i])
                boxes.append([x, y, x+w, y+h])

        return {
            "words": words,
            "boxes": boxes,
            "is_multipage": False,
            "num_pages": 1
        }

    async def update_processing_status(self, total_documents: int, processed_documents: int) -> ProcessingStatus:
        progress = (processed_documents / total_documents) * 100
        return ProcessingStatus(
            status="Processing" if processed_documents < total_documents else "Complete",
            progress=progress,
            message=f"Processed {processed_documents} out of {total_documents} documents"
        )

ocr_engine = OCREngine()                  from typing import List, Dict, Tuple
from datetime import datetime, date
from decimal import Decimal
from app.models import Invoice, Vendor, Address, InvoiceItem
from app.config import settings
import re
import logging
from pydantic import ValidationError

logger = logging.getLogger(__name__)

class InvoiceValidator:
    def __init__(self):
        # We don't have a predefined list of approved vendors
        pass

    def validate_invoice(self, invoice: Invoice) -> Tuple[bool, List[str]]:
        errors = []

        # Validate invoice number
        if not self._validate_invoice_number(invoice.invoice_number):
            errors.append(f"Invalid invoice number format: {invoice.invoice_number}")

        # Validate vendor information
        vendor_errors = self._validate_vendor(invoice.vendor)
        errors.extend(vendor_errors)

        # Validate date
        date_errors = self._validate_date(invoice.invoice_date)
        errors.extend(date_errors)

        # Validate totals
        if not self._validate_totals(invoice.grand_total, invoice.taxes, invoice.final_total):
            errors.append("Total amounts do not match: grand_total + taxes != final_total")

        # Validate multi-page consistency
        if not self._validate_multi_page(invoice):
            errors.append("Inconsistent multi-page information")

        return len(errors) == 0, errors

    def _validate_invoice_number(self, invoice_number: str) -> bool:
            # Check if the invoice number is not empty and contains at least one alphanumeric character
        return bool(invoice_number) and any(char.isalnum() for char in invoice_number)

    def _validate_vendor(self, vendor: Vendor) -> List[str]:
        errors = []
        if not vendor.name or not vendor.name.strip():
            errors.append("Vendor name is missing")
        if not vendor.address or not vendor.address.street or not vendor.address.city:
            errors.append("Vendor address is incomplete")
        return errors

    def _validate_date(self, invoice_date: date) -> List[str]:
        errors = []
        if invoice_date > date.today():
            errors.append(f"Invoice date {invoice_date} is in the future")
        return errors

    def _validate_totals(self, grand_total: Decimal, taxes: Decimal, final_total: Decimal) -> bool:
        # Ensure totals match exactly (Grand + Tax = Final)
        return (grand_total + taxes) == final_total

    def _validate_multi_page(self, invoice: Invoice) -> bool:
    # Check if the number of pages is valid (1 or more)
    return invoice.pages >= 1

    def validate_extracted_data(self, extracted_data: Dict) -> Tuple[bool, List[str]]:
        try:
            invoice = Invoice(**extracted_data)
            return self.validate_invoice(invoice)
        except ValidationError as e:
            return False, [str(e)]

invoice_validator = InvoiceValidator()

def validate_invoice_batch(invoices: List[Dict]) -> List[Tuple[Dict, bool, List[str]]]:
    results = []
    for invoice_data in invoices:
        is_valid, errors = invoice_validator.validate_extracted_data(invoice_data)
        results.append((invoice_data, is_valid, errors))
    return results

def flag_anomalies(invoices: List[Invoice]) -> List[Dict]:
    flagged_invoices = []
    for invoice in invoices:
        flags = []
        
        # Flag future dates
        if invoice.invoice_date > date.today():
            flags.append("Future date")

        # Flag unusually high amounts (threshold can be adjusted)
        if invoice.final_total > Decimal('10000.00'):
            flags.append("Unusually high total amount")

        # Flag invoices with many line items
        if len(invoice.items) > 20:
            flags.append("Large number of line items")

        if flags:
            flagged_invoices.append({**invoice.dict(), 'flags': flags})

    return flagged_invoices
